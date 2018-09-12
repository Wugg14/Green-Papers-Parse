import openpyxl, datetime, csv

print('Opening workbooks')

wb1 = openpyxl.load_workbook('CandidateSummer2018.xlsx')

wb2 = openpyxl.load_workbook('NewSheet.xlsx')


sheet1 = wb1['CandidateSummer2018']
sheet2 = wb2['Sheet1']

counter = 0
counter2 = 0

def nameChecker(lastName1, lastName2, firstInitial1, firstInitial2, row1, row2):
    lastName1 = lastName1.lower()
    lastName1 = ''.join(e for e in lastName1 if e.isalnum())
    lastName1 = lastName1[:3]
    lastName2 = lastName2.lower()
    lastName2 = ''.join(e for e in lastName2 if e.isalnum())
    lastName2 = lastName2[:3]
    firstInitial1 = firstInitial1.lower()
    firstInitial2 = firstInitial2.lower()
    if lastName1 == lastName2:
        if firstInitial1 == firstInitial2:
            global counter
            counter += 1
            print(firstInitial1 + ' ' + lastName1 + ' ' + str(counter))
            #Add the Email
            if sheet2['K' + str(row2)].value != '':
                sheet1['F' + str(row1)].value = sheet2['K' + str(row2)].value
                global counter2
                counter2 += 1
                print(counter2)
            #Add the Phone Number
            if sheet2['M' + str(row2)].value != '':
                sheet1['G' + str(row1)].value = sheet2['M' + str(row2)].value
            #Check to see it theyve already signed the pledge
            if row2 < 172:
                sheet1['H' + str(row1)].value = 'Yes'

#loop through spreadsheets
for row1 in range(2, sheet1.max_row + 1):
    for row2 in range(2, sheet2.max_row + 1):
        #name checking variables
        lastName1 = sheet1['D' + str(row1)].value
        lastName2 = sheet2['G' + str(row2)].value
        firstInitial1 = sheet1['C' + str(row1)].value[0]
        firstInitial2 = sheet2['H' + str(row2)].value[0]
        #Check if state abbreviations are the same
        if  sheet1['A' + str(row1)].value == sheet2['A' + str(row2)].value:
            #create a variable for the race/district code
            district1 = sheet1['B'+ str(row1)].value
            district2 = sheet2['C' + str(row2)].value
            district2 = district2.lower()
            #if both are the same senate race
            if district1[0] == 'S' and district2 == 'senate':
                nameChecker(lastName1, lastName2, firstInitial1, firstInitial2, row1, row2)
            #check if house race
            elif district1[0] == 'H' and district2 =='house':
                #check house race district
                district3 = sheet2['B' + str(row2)].value
                district4 = sheet1['B'+ str(row1)].value[1:]
                if district4[0] == '0':
                    district4 = district4[-1]
                if district4 == str(district3):
                    nameChecker(lastName1, lastName2, firstInitial1, firstInitial2, row1, row2)

wb1.save('CandidateSummer2018.xlsx')

   
